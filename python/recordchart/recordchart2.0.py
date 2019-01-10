
# Author: tHaninG
# Email: tHan.ninG0@outlook.com
# Date:   2017-10-22 21:55:00
# Last Modified time: 2017-11-06 14:19:26
# -*- coding: utf-8 -*-

print('作者: tHaninG\n邮箱: tHan.ninG@outlook.com\n有任何问题请联系我. \n载入中......')

from os import system
from tkinter import Tk
from tkinter import filedialog
from openpyxl import load_workbook

# set time format Y-M-D-T
time_format = '%Y-%m-%d'

# initialize Tk()
root = Tk()
root.withdraw()

# get chart's file path
system("cls")
input('按 Enter 选择文件\n')
fpath = filedialog.askopenfilename()
record = load_workbook(fpath)  # load xlsx
sheet_name = record.get_sheet_names()  # get sheet names


n_sheet = len(sheet_name)  # number of sheets

if n_sheet == 1:
    sheet = record.get_sheet_by_name(sheet_name[0])  # load sheet
else:
    option = []
    brace1 = ''
    brace2 = ''
    for i in range(1, n_sheet + 1):
        option.append(chr(64 + i))
        # left-justified
        brace1 = brace1 + ' {}'.ljust(2 + len(sheet_name[i - 1]))
        brace2 = brace2 + ' {}'
    while(1):
        print('\n' + brace1[1:].format(*option) +
              '\n' + brace2[1:].format(*sheet_name))
        chose_sheet = str.upper(input('输入字母选择表格:\n'))
        system("cls")
        if chose_sheet in option:
            sheet = record.get_sheet_by_name(
                sheet_name[ord(chose_sheet) - 65])  # lshee
            break
        else:
            print('\nERROR:Invalid Input\n请重试')


# Check all rows and update given ID's record
def update(ID, c_IDname, c_time, c_name, sheet):
    """
    Update record of student ID

    Arguments:
    ID -- student ID ('2017222010032')
    c_IDname -- the column name of ID ('B')
    c_time -- the column name of times ('J')
    c_name -- the column name of student name
    sheet -- data sheet

    Return:
    0/1 -- whether update success or not
    """
    n = 0
    backup = []
    produceID = None

    # find the position of ID
    for rowNum in range(2, sheet.max_row):
        produceID = sheet[c_IDname + str(rowNum)].value
        if str(produceID).find(ID) >= 0:
            n = n + 1
            # ensure the ID is unique
            if n == 1:
                sheet[c_time + str(rowNum)].value = chr(10003)  # check
                backup = str(rowNum)
            else:
                sheet[c_time + backup].value = None
                break

    if n == 0:
        print('\nERROR:can\'t find this ID!!! \n请重试')
        return 0

    elif n == 1:
        print('\n%s 更新成功' %
              sheet[c_name + backup].value)

        # prevent wrong ID
        while (1):
            opt = input(
                '\n按 Enter 确认\n(输入 C 取消)\n')
            if opt == 'C':
                sheet[c_time + backup].value = None
                print('\n更新取消')
                return 0
            elif opt == '':
                system("cls")
                return 1
            else:
                print('\nERROR:Invalid Input\n请重试')

    else:
        print('\nERROR:ID is not unique!!! \n按 Enter 重试')
        return 0


# Check all rows and update given ID's record and grade
def update_grade(ID, c_IDname, c_time, c_name, sheet):
    """
    Update grade of student ID

    Arguments:
    ID -- student ID ('2017222010032')
    c_IDname -- the column name of ID ('B')
    c_time -- the column name of times ('J')
    c_name -- the column name of student name
    sheet -- data sheet

    Return:
    0/1 -- whether update success or not
    """
    n = 0
    backup = []
    produceID = None

    # find the position of ID
    for rowNum in range(2, sheet.max_row):
        produceID = sheet[c_IDname + str(rowNum)].value
        if str(produceID).find(ID) >= 0:
            n = n + 1
            # ensure the ID is unique
            if n == 1:
                backup = str(rowNum)
            else:
                break
    if n == 0:
        print('\nERROR:can\'t find this ID!!! \n请重试')
        return 0

    elif n == 1:
        while (1):
            print('\n请输入 %s 的成绩:' %
                  sheet[c_name + backup].value)
            print('(输入 C 取消)')
            grade_input = input()
            try:
                grade_bu = int(grade_input)
                if grade_bu >= 0 and grade_bu <= 100:
                    sheet[c_time + str(backup)].value = str(grade_bu)
                    system("cls")
                    print('%s 更新成功\n' %
                          sheet[c_name + backup].value)
                    return 1
                else:
                    print('\nERROR:Invalid Input\n请重试')
            except:
                if grade_input == 'C':
                    print('\n更新取消')
                    return 0
                else:
                    print('\nERROR:Invalid Input\n请重试')

    else:
        print('\nERROR:ID is not unique!!! \n请重试')
        return 0

# Check all rows and update given ID's record and rank


def update_rank(ID, c_IDname, c_time, c_name, sheet):
    """
    Update rank of student ID

    Arguments:
    ID -- student ID ('2017222010032')
    c_IDname -- the column name of ID ('B')
    c_time -- the column name of times ('J')
    c_name -- the column name of student name
    sheet -- data sheet

    Return:
    0/1 -- whether update success or not
    """
    n = 0
    backup = []
    produceID = None

    # find the position of ID
    for rowNum in range(2, sheet.max_row):
        produceID = sheet[c_IDname + str(rowNum)].value
        if str(produceID).find(ID) >= 0:
            n = n + 1
            # ensure the ID is unique
            if n == 1:
                backup = str(rowNum)
            else:
                break
    if n == 0:
        print('\nERROR:can\'t find this ID!!! \n请重试')
        return 0

    elif n == 1:
        while (1):
            print('\n请输入 %s 的等级:' %
                  sheet[c_name + backup].value)
            print('(输入 CC 取消)')
            rank_input = str.upper(input())

            if 'ABCD+-'.find(rank_input) >= 0:
                sheet[c_time + str(backup)].value = rank_input
                system("cls")
                print('%s 更新成功\n' %
                      sheet[c_name + backup].value)
                return 1
            elif rank_input=='CC':
                return 0
            else:
                print('\nERROR:Invalid Input\n请重试')

    else:
        print('\nERROR:ID is not unique!!! \n请重试')
        return 0


# chose record what
while(1):
    while(1):
        rg = str.upper(input('A:成绩 B:等级(A,B,C) C:签到\n选择任务:\n'))
        if 'ABC'.find(rg)>=0:
            opt = input('\n任务为 '+str(rg)+'\n按 Enter 确定\n(输入 \'CC\' 取消)\n')
            if opt=='':
                system("cls")
                break
            elif opt=='CC':
                pass
            else:
                print('\nERROR:Invalid Input\n请重试')
        else:
            print('\nERROR:Invalid Input\n请重试')

    # input index of ID,name,...
    index_name = ['学号列号',
                  '时间列号', '姓名列号']
    index = ['','','']
    num = -1
    for i in index_name:
        num = num + 1
        while(1):
            index[num]=str.upper(input('输入 ' + str(i) + ':\n'))
            while(1):
                opt = input('\n按 Enter 确定\n(输入 \'CC\' 取消)\n')
                if opt=='' or opt=='CC':
                    break
                else:
                    print('\nERROR:Invalid Input\n请重试')
            if opt=='':
                system("cls")
                break


    # record check
    if rg == 'C':
        # update all data
        k = 0
        while(1):
            # input student ID
            ID = str(
                input('请输入 学号:\n(输入 stop 退出更新)\n'))
            if ID == 'stop':
                system("cls")
                break

            # update data
            k = k + update(ID, index[0], index[1], index[2], sheet)
        break

    # record grade
    elif rg == 'A':
        # update all data
        k = 0
        while(1):
            # input student ID
            ID = str(
                input('请输入 学号:\n(输入 stop 退出更新)\n'))
            if ID == 'stop':
                system("cls")
                break

            # update data
            k = k + update_grade(ID, index[0], index[1], index[2], sheet)
        break

    # record rank
    elif rg == 'B':
        # update all data
        k = 0
        while(1):
            # input student ID
            ID = str(
                input('请输入 学号:\n(输入 stop 退出更新)\n'))
            if ID == 'stop':
                system("cls")
                break

            # update data
            k = k + update_rank(ID, index[0], index[1], index[2], sheet)
        break

    else:
        print('\nERROR:Invalid Input\n请重试')


# save updated recrodchart
while(1):
    sv = str.upper(input('是否保存? Y/N\n'))
    if sv == 'Y':
        while(1):
            cover = str.upper(input(
                '\n是否覆盖原文件? Y/N\n(输入 \'N\' 将在原文件的路径保存一个备份)\n'))
            if cover == 'Y':
                try:
                    record.save(fpath)
                except:
                    record.save(fpath[:-5] + '(备份)' + '.xlsx')
                    print('\n无法覆盖原文件\n已保存为备份')
                break
            elif cover == 'N':
                record.save(fpath[:-5] + '(备份)' + '.xlsx')
                break
            else:
                print('\nERROR:Invalid Input\n请重试')

        system("cls")
        print('任务完成!!!' +
              '\n更新数据总数: ' + str(k))
        input('\n按 Enter 退出\n')
        break

    elif sv == 'N':
        system("cls")
        print('任务取消!!!')
        input('\n按 Enter 退出\n')
        break
    else:
        print('\nERROR:Invalid Input\n请重试\n')
